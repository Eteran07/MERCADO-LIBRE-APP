import tkinter as tk
from tkinter import ttk, filedialog, messagebox
import openpyxl
from openpyxl.styles import PatternFill, Font, Border, Side, Alignment
import shutil
import os
import threading
import difflib

# --- CONFIGURACIÓN: HISTORIAL DE CAMBIOS ---
CHANGELOG_TEXT = """
HISTORIAL DE VERSIONES
==================================================

[Versión 6.3] - ACTUAL (REPORT GUARD)
--------------------------------------------------
• Reportes: Estructura de reportes estandarizada para evitar columnas desorganizadas.
• Interfaz: Nueva ventana de "Vista Previa de Reporte" antes de guardar, para que puedas validar los datos.
• Estabilidad: Mejor manejo de celdas combinadas y formatos extraños al leer.

[Versión 6.2]
--------------------------------------------------
• Estabilidad: Escritura "Quirúrgica".
• Visual: Reportes estilizados.

[Versión 6.1]
--------------------------------------------------
• Mapeo Independiente: Configuración por hoja.
"""

# --- COLORES Y ESTILOS UI ---
COLOR_BG_MAIN = "#f4f6f9"
COLOR_BG_CARD = "#ffffff"
COLOR_PRIMARY = "#00a650"
COLOR_TEXT_HEADER = "#2d3436"
COLOR_ACCENT = "#3498db"

# --- UTILIDADES DE ESTILO EXCEL ---
def estilizar_hoja_excel(ws):
    """Aplica formato profesional a una hoja de Excel generada"""
    header_fill = PatternFill(start_color="3498db", end_color="3498db", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), 
                         top=Side(style='thin'), bottom=Side(style='thin'))
    
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = thin_border

    for column_cells in ws.columns:
        length = max(len(str(cell.value)) if cell.value else 0 for cell in column_cells)
        ws.column_dimensions[column_cells[0].column_letter].width = min(length + 4, 60)

    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.border = thin_border

class ReportePreviewDialog(tk.Toplevel):
    """Ventana para previsualizar datos antes de generar reporte"""
    def __init__(self, parent, title, data, columns):
        super().__init__(parent)
        self.title(f"Vista Previa: {title}")
        self.geometry("800x500")
        self.confirmado = False
        
        ttk.Label(self, text=f"Confirmar datos para: {title}", font=("Segoe UI", 12, "bold")).pack(pady=10)
        
        # Tabla
        frame_table = ttk.Frame(self)
        frame_table.pack(fill=tk.BOTH, expand=True, padx=20, pady=5)
        
        tree = ttk.Treeview(frame_table, columns=columns, show="headings", height=15)
        scroll_y = ttk.Scrollbar(frame_table, orient="vertical", command=tree.yview)
        scroll_x = ttk.Scrollbar(frame_table, orient="horizontal", command=tree.xview)
        tree.configure(yscrollcommand=scroll_y.set, xscrollcommand=scroll_x.set)
        
        tree.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        scroll_y.pack(side=tk.RIGHT, fill=tk.Y)
        scroll_x.pack(side=tk.BOTTOM, fill=tk.X)
        
        for col in columns:
            tree.heading(col, text=col)
            tree.column(col, width=120)
            
        # Insertar primeros 50 registros para preview
        for item in data[:50]:
            tree.insert("", "end", values=item)
            
        ttk.Label(self, text=f"Mostrando primeros {min(len(data), 50)} de {len(data)} registros.", foreground="gray").pack(pady=5)
        
        frame_btns = ttk.Frame(self)
        frame_btns.pack(pady=15)
        ttk.Button(frame_btns, text="Cancelar Reporte", command=self.destroy).pack(side=tk.LEFT, padx=10)
        ttk.Button(frame_btns, text="✅ Aprobar y Generar", command=self.aprobar, style="Action.TButton").pack(side=tk.LEFT, padx=10)

    def aprobar(self):
        self.confirmado = True
        self.destroy()

class ExcelPreviewFrame(ttk.Frame):
    """Componente para visualizar y configurar Excel"""
    def __init__(self, parent, title, is_multisheet=False, on_sheet_change=None):
        super().__init__(parent, style="Card.TFrame", padding=15)
        self.file_path = None
        self.sheet_data = [] 
        self.headers = []
        self.is_multisheet = is_multisheet
        self.on_sheet_change_callback = on_sheet_change
        
        lbl_title = ttk.Label(self, text=title, style="CardTitle.TLabel")
        lbl_title.pack(anchor="w", pady=(0, 10))

        f_top = ttk.Frame(self, style="Card.TFrame")
        f_top.pack(fill=tk.X, pady=5)
        
        self.btn_load = ttk.Button(f_top, text="📂 Cargar Archivo", command=self.cargar_archivo, style="Outline.TButton")
        self.btn_load.pack(side=tk.LEFT)
        
        self.lbl_file = ttk.Label(f_top, text="Sin archivo seleccionado", foreground="#95a5a6", style="Card.TLabel")
        self.lbl_file.pack(side=tk.LEFT, padx=10)

        f_mid = ttk.Frame(self, style="Card.TFrame")
        f_mid.pack(fill=tk.X, pady=10)

        if self.is_multisheet:
            ttk.Label(f_mid, text="Selecciona las Hojas a procesar:", style="Bold.TLabel").pack(anchor="w")
            f_list = ttk.Frame(f_mid)
            f_list.pack(fill=tk.X, pady=5)
            
            self.listbox_sheets = tk.Listbox(f_list, selectmode=tk.MULTIPLE, height=4, exportselection=False)
            scroll_list = ttk.Scrollbar(f_list, orient="vertical", command=self.listbox_sheets.yview)
            self.listbox_sheets.config(yscrollcommand=scroll_list.set)
            
            self.listbox_sheets.pack(side=tk.LEFT, fill=tk.X, expand=True)
            scroll_list.pack(side=tk.LEFT, fill=tk.Y)
            
            f_btns = ttk.Frame(f_mid)
            f_btns.pack(fill=tk.X)
            ttk.Button(f_btns, text="Seleccionar Todas", command=self.seleccionar_todas, style="Small.TButton").pack(side=tk.LEFT)
            
        else:
            ttk.Label(f_mid, text="Hoja de trabajo:", style="Card.TLabel").pack(side=tk.LEFT)
            self.combo_sheet = ttk.Combobox(f_mid, state="disabled", width=30)
            self.combo_sheet.pack(side=tk.LEFT, padx=10)
            self.combo_sheet.bind("<<ComboboxSelected>>", self.cambiar_hoja_combo)

        if not self.is_multisheet:
            f_table = ttk.Frame(self, style="Card.TFrame")
            f_table.pack(fill=tk.BOTH, expand=True, pady=10)
            
            scroll_y = ttk.Scrollbar(f_table, orient="vertical")
            scroll_x = ttk.Scrollbar(f_table, orient="horizontal")
            
            self.tree = ttk.Treeview(f_table, columns=[], show="headings", height=6, 
                                     yscrollcommand=scroll_y.set, xscrollcommand=scroll_x.set,
                                     style="Modern.Treeview")
            
            scroll_y.config(command=self.tree.yview)
            scroll_x.config(command=self.tree.xview)
            
            scroll_y.pack(side=tk.RIGHT, fill=tk.Y)
            scroll_x.pack(side=tk.BOTTOM, fill=tk.X)
            self.tree.pack(fill=tk.BOTH, expand=True)
        
        self.frame_cols = ttk.Frame(self, style="Card.TFrame")
        self.frame_cols.pack(fill=tk.X, pady=5)

    def cargar_archivo(self):
        filename = filedialog.askopenfilename(filetypes=[("Excel Files", "*.xlsx *.xls")])
        if filename:
            try:
                f = open(filename, "r+")
                f.close()
            except PermissionError:
                messagebox.showerror("Error de Archivo", f"El archivo está ABIERTO.\nPor favor ciérralo antes de cargarlo:\n{os.path.basename(filename)}")
                return
            except: pass

            self.file_path = filename
            self.lbl_file.config(text=os.path.basename(filename), foreground=COLOR_PRIMARY)
            try:
                wb = openpyxl.load_workbook(filename, read_only=True, data_only=True)
                sheets = wb.sheetnames
                wb.close()

                if self.is_multisheet:
                    self.listbox_sheets.delete(0, tk.END)
                    for sh in sheets:
                        self.listbox_sheets.insert(tk.END, sh)
                    self.listbox_sheets.select_set(0)
                else:
                    self.combo_sheet.config(values=sheets, state="readonly")
                    default = "Publicaciones" if "Publicaciones" in sheets else sheets[0]
                    self.combo_sheet.set(default)
                    self.cargar_preview_hoja(default)
                    
            except Exception as e:
                messagebox.showerror("Error", f"No se pudo leer:\n{e}")

    def seleccionar_todas(self):
        if self.is_multisheet:
            self.listbox_sheets.select_set(0, tk.END)

    def cambiar_hoja_combo(self, event):
        hoja = self.combo_sheet.get()
        self.cargar_preview_hoja(hoja)

    def cargar_preview_hoja(self, hoja_nombre):
        if not self.file_path or self.is_multisheet: return
        try:
            wb = openpyxl.load_workbook(self.file_path, read_only=True, data_only=True)
            ws = wb[hoja_nombre]
            raw_rows = []
            for i, row in enumerate(ws.iter_rows(values_only=True)):
                if i >= 20: break
                raw_rows.append(list(row))
            wb.close()
            
            header_row_idx = 0
            for i, row in enumerate(raw_rows):
                row_str = " ".join([str(x).upper() if x is not None else "" for x in row])
                if "SKU" in row_str or "CODIGO" in row_str or "PRICE" in row_str or "PRECIO" in row_str:
                    header_row_idx = i
                    break
            
            if raw_rows:
                self.headers = raw_rows[header_row_idx]
                self.sheet_data = raw_rows[header_row_idx+1:]
            else:
                self.headers = []
                self.sheet_data = []

            self.actualizar_tabla()
            
            if self.on_sheet_change_callback:
                cols_procesadas = []
                for i, h in enumerate(self.headers):
                    val = str(h).strip() if h is not None else ""
                    if val == "": val = f"Columna {i+1}"
                    cols_procesadas.append(val)
                self.on_sheet_change_callback(cols_procesadas)
                
        except Exception as e:
            print(f"Error preview: {e}")

    def actualizar_tabla(self):
        if self.is_multisheet: return
        self.tree.delete(*self.tree.get_children())
        display_cols = []
        col_ids = []
        for i, h in enumerate(self.headers):
            val = str(h).strip() if h is not None else ""
            col_id = f"col_{i}"
            col_ids.append(col_id)
            if val == "": display_cols.append(f"Columna {i+1}")
            else: display_cols.append(val)

        self.tree["columns"] = col_ids
        for i, col_id in enumerate(col_ids):
            self.tree.heading(col_id, text=display_cols[i])
            self.tree.column(col_id, width=120, minwidth=50) 
            
        for index, row in enumerate(self.sheet_data):
            clean_row = []
            for i in range(len(col_ids)):
                val = ""
                if i < len(row):
                    val = str(row[i]) if row[i] is not None else ""
                clean_row.append(val)
            tag = 'even' if index % 2 == 0 else 'odd'
            self.tree.insert("", "end", values=clean_row, tags=(tag,))
            
    def get_selected_sheets(self):
        if self.is_multisheet:
            indices = self.listbox_sheets.curselection()
            return [self.listbox_sheets.get(i) for i in indices]
        else:
            return [self.combo_sheet.get()]


class ConfigHojaDialog(tk.Toplevel):
    def __init__(self, parent, file_path, sheet_name):
        super().__init__(parent)
        self.title(f"Configurar: {sheet_name}")
        self.geometry("900x600")
        self.result = None
        self.file_path = file_path
        self.sheet_name = sheet_name
        self.headers = []
        
        style = ttk.Style()
        style.configure("Dialog.TLabel", font=("Segoe UI", 10))
        
        ttk.Label(self, text=f"Hoja: '{sheet_name}'", font=("Segoe UI", 14, "bold")).pack(pady=10)
        ttk.Label(self, text="Confirma las columnas para esta hoja.", style="Dialog.TLabel").pack()

        f_table = ttk.Frame(self)
        f_table.pack(fill=tk.BOTH, expand=True, padx=20, pady=10)
        
        scroll_y = ttk.Scrollbar(f_table, orient="vertical")
        scroll_x = ttk.Scrollbar(f_table, orient="horizontal")
        
        self.tree = ttk.Treeview(f_table, columns=[], show="headings", height=10, 
                                 yscrollcommand=scroll_y.set, xscrollcommand=scroll_x.set)
        
        scroll_y.config(command=self.tree.yview)
        scroll_x.config(command=self.tree.xview)
        scroll_y.pack(side=tk.RIGHT, fill=tk.Y)
        scroll_x.pack(side=tk.BOTTOM, fill=tk.X)
        self.tree.pack(fill=tk.BOTH, expand=True)

        f_cols = ttk.Frame(self)
        f_cols.pack(fill=tk.X, padx=20, pady=10)
        
        self.var_sku = tk.StringVar()
        self.var_price = tk.StringVar()
        self.var_name = tk.StringVar()
        
        ttk.Label(f_cols, text="Columna SKU:", font=("Segoe UI", 9, "bold")).grid(row=0, column=0, padx=5)
        self.cb_sku = ttk.Combobox(f_cols, textvariable=self.var_sku, state="readonly", width=25)
        self.cb_sku.grid(row=1, column=0, padx=5)
        
        ttk.Label(f_cols, text="Columna PRECIO:", font=("Segoe UI", 9, "bold")).grid(row=0, column=1, padx=5)
        self.cb_price = ttk.Combobox(f_cols, textvariable=self.var_price, state="readonly", width=25)
        self.cb_price.grid(row=1, column=1, padx=5)
        
        ttk.Label(f_cols, text="Columna NOMBRE (Opcional):", font=("Segoe UI", 9, "bold")).grid(row=0, column=2, padx=5)
        self.cb_name = ttk.Combobox(f_cols, textvariable=self.var_name, state="readonly", width=25)
        self.cb_name.grid(row=1, column=2, padx=5)

        ttk.Button(self, text="✅ Confirmar y Siguiente", command=self.confirmar, style="Action.TButton").pack(pady=20)
        self.cargar_datos()

    def cargar_datos(self):
        try:
            wb = openpyxl.load_workbook(self.file_path, read_only=True, data_only=True)
            ws = wb[self.sheet_name]
            raw_rows = []
            for i, row in enumerate(ws.iter_rows(values_only=True)):
                if i >= 15: break
                raw_rows.append(list(row))
            wb.close()
            
            header_row_idx = 0
            for i, row in enumerate(raw_rows):
                row_str = " ".join([str(x).upper() if x is not None else "" for x in row])
                if "SKU" in row_str or "CODIGO" in row_str or "PRICE" in row_str or "PRECIO" in row_str:
                    header_row_idx = i
                    break
            
            self.headers = []
            if raw_rows:
                raw_headers = raw_rows[header_row_idx]
                sheet_data = raw_rows[header_row_idx+1:]
                
                self.headers = []
                for i, h in enumerate(raw_headers):
                    val = str(h).strip() if h is not None else ""
                    clean_val = val if val else f"Columna {i+1}"
                    self.headers.append(clean_val)
                
                self.tree["columns"] = self.headers
                for col in self.headers:
                    self.tree.heading(col, text=col)
                    self.tree.column(col, width=100)
                
                for row in sheet_data:
                    clean_row = [str(x) if x is not None else "" for x in row]
                    while len(clean_row) < len(self.headers): clean_row.append("")
                    self.tree.insert("", "end", values=clean_row[:len(self.headers)])
                
                self.cb_sku.config(values=self.headers)
                self.cb_price.config(values=self.headers)
                self.cb_name.config(values=self.headers)
                
                for h in self.headers:
                    hu = h.upper()
                    if any(k in hu for k in ["SKU", "CODIGO", "ZMART"]) and not self.var_sku.get(): self.var_sku.set(h)
                    if any(k in hu for k in ["PRECIO", "COSTO", "ML"]) and not self.var_price.get(): self.var_price.set(h)
                    if any(k in hu for k in ["PRODUCTO", "NOMBRE"]) and not self.var_name.get(): self.var_name.set(h)

        except Exception as e:
            messagebox.showerror("Error", str(e))
            self.destroy()

    def confirmar(self):
        if not self.var_sku.get() or not self.var_price.get():
            messagebox.showwarning("Faltan datos", "Debes seleccionar al menos SKU y Precio.")
            return
        self.result = {"sku": self.var_sku.get(), "price": self.var_price.get(), "name": self.var_name.get()}
        self.destroy()


class ActualizadorApp:
    def __init__(self, root):
        self.root = root
        self.root.title("Actualizador ML v6.3 (Report Guard)")
        self.root.geometry("1100x950")
        self.root.configure(bg=COLOR_BG_MAIN)
        try:
            if os.path.exists("logo.ico"): self.root.iconbitmap("logo.ico")
        except: pass
        self.setup_styles()
        
        main_frame = tk.Frame(root, bg=COLOR_BG_MAIN)
        main_frame.pack(fill=tk.BOTH, expand=True)
        self.canvas = tk.Canvas(main_frame, bg=COLOR_BG_MAIN, highlightthickness=0)
        scrollbar_y = ttk.Scrollbar(main_frame, orient="vertical", command=self.canvas.yview)
        
        self.scrollable_frame = tk.Frame(self.canvas, bg=COLOR_BG_MAIN)
        self.scrollable_frame.bind("<Configure>", lambda e: self.canvas.configure(scrollregion=self.canvas.bbox("all")))
        self.canvas_window = self.canvas.create_window((0, 0), window=self.scrollable_frame, anchor="nw")
        self.canvas.configure(yscrollcommand=scrollbar_y.set)
        self.canvas.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        scrollbar_y.pack(side=tk.RIGHT, fill=tk.Y)
        self.canvas.bind("<Configure>", self.on_canvas_configure)

        # UI Components
        header = tk.Frame(self.scrollable_frame, bg="white", height=80, padx=20, pady=15)
        header.pack(fill=tk.X, pady=(0, 20))
        tk.Frame(header, bg="#e0e0e0", height=1).place(relx=0, rely=1, relwidth=1, anchor="sw")
        ttk.Label(header, text="🛠️ Consola de Actualización", style="BigHeader.TLabel").pack(side=tk.LEFT)
        ttk.Button(header, text="ℹ  Ver Historial", style="Info.TButton", command=self.mostrar_changelog).pack(side=tk.RIGHT)

        content = tk.Frame(self.scrollable_frame, bg=COLOR_BG_MAIN, padx=30)
        content.pack(fill=tk.BOTH, expand=True)

        self.frame_origen = ExcelPreviewFrame(content, "1. ORIGEN: Matriz de Precios (Multi-Hojas)", is_multisheet=True)
        self.frame_origen.pack(fill=tk.X, pady=10)
        ttk.Label(self.frame_origen.frame_cols, text="ℹ Al iniciar, se te pedirá confirmar las columnas para CADA hoja seleccionada.", style="Dim.TLabel").pack()

        self.frame_destino = ExcelPreviewFrame(content, "2. DESTINO: Mercado Libre", is_multisheet=False, on_sheet_change=self.cols_destino_cargadas)
        self.frame_destino.pack(fill=tk.X, pady=20)
        
        f_dst = self.frame_destino.frame_cols
        l4 = ttk.Label(f_dst, text="Columna SKU (ML):", style="Bold.TLabel")
        l4.grid(row=0, column=0, sticky="w", padx=5)
        self.cb_sku_dest = ttk.Combobox(f_dst, state="disabled", width=35)
        self.cb_sku_dest.grid(row=1, column=0, padx=5, pady=5)
        
        l5 = ttk.Label(f_dst, text="Columna TÍTULO (ML):", style="Bold.TLabel")
        l5.grid(row=0, column=1, sticky="w", padx=5)
        self.cb_title_dest = ttk.Combobox(f_dst, state="disabled", width=35)
        self.cb_title_dest.grid(row=1, column=1, padx=5, pady=5)

        action_frame = tk.Frame(content, bg=COLOR_BG_MAIN)
        action_frame.pack(fill=tk.X, pady=20)
        self.progress_var = tk.DoubleVar()
        self.progress_bar = ttk.Progressbar(action_frame, variable=self.progress_var, maximum=100)
        self.progress_bar.pack(fill=tk.X, pady=(0, 10))
        self.lbl_progress = ttk.Label(action_frame, text="Listo.", style="Dim.TLabel")
        self.lbl_progress.pack(anchor="center")

        self.btn_run = ttk.Button(action_frame, text="🚀  CONFIGURAR Y PROCESAR", command=self.iniciar_flujo, style="Action.TButton")
        self.btn_run.pack(fill=tk.X, pady=10, ipady=8)
        self.btn_run.state(["disabled"])

        ttk.Label(content, text="Registro de actividades:", style="Bold.TLabel").pack(anchor="w")
        self.txt_log = tk.Text(content, height=8, bg="#2d3436", fg="#00cec9", font=("Consolas", 10), relief="flat", padx=10, pady=10)
        self.txt_log.pack(fill=tk.X, pady=(5, 30))

    def setup_styles(self):
        style = ttk.Style()
        style.theme_use('clam')
        style.configure("TFrame", background=COLOR_BG_MAIN)
        style.configure("Card.TFrame", background=COLOR_BG_CARD, relief="flat", borderwidth=1)
        style.configure("BigHeader.TLabel", font=("Segoe UI", 18, "bold"), foreground=COLOR_TEXT_HEADER, background="white")
        style.configure("Card.TLabel", font=("Segoe UI", 10), background=COLOR_BG_CARD)
        style.configure("CardTitle.TLabel", font=("Segoe UI", 12, "bold"), foreground=COLOR_PRIMARY, background=COLOR_BG_CARD)
        style.configure("Bold.TLabel", font=("Segoe UI", 10, "bold"), background=COLOR_BG_CARD)
        style.configure("Dim.TLabel", font=("Segoe UI", 9), foreground="#7f8c8d", background=COLOR_BG_MAIN)
        style.configure("Outline.TButton", font=("Segoe UI", 9), background="#ecf0f1")
        style.configure("Small.TButton", font=("Segoe UI", 8), background="#dfe6e9")
        style.configure("Action.TButton", font=("Segoe UI", 12, "bold"), background=COLOR_PRIMARY, foreground="white", borderwidth=0)
        style.map("Action.TButton", background=[('active', '#008f39'), ('disabled', '#bdc3c7')])
        style.configure("Info.TButton", font=("Segoe UI", 9), background=COLOR_ACCENT, foreground="white")
        style.configure("Modern.Treeview", background="white", fieldbackground="white", rowheight=25, borderwidth=0)
        style.configure("Modern.Treeview.Heading", font=("Segoe UI", 9, "bold"), background="#ecf0f1", relief="flat")
        style.map("Modern.Treeview.Heading", background=[('active', '#bdc3c7')])

    def on_canvas_configure(self, event):
        self.canvas.itemconfig(self.canvas_window, width=event.width)

    def mostrar_changelog(self):
        vent_info = tk.Toplevel(self.root)
        vent_info.title("Historial")
        vent_info.geometry("600x500")
        try:
            if os.path.exists("logo.ico"): vent_info.iconbitmap("logo.ico")
        except: pass
        txt = tk.Text(vent_info, font=("Consolas", 10), padx=20, pady=20, bg="white", relief="flat")
        txt.pack(fill=tk.BOTH, expand=True)
        txt.insert(tk.END, CHANGELOG_TEXT)
        txt.config(state="disabled")

    def cols_destino_cargadas(self, columnas):
        self.cb_sku_dest.config(values=columnas, state="readonly")
        self.cb_title_dest.config(values=columnas, state="readonly")
        for c in columnas:
            val = str(c).strip()
            if val in ["SELLER_SKU", "SKU", "CODIGO", "INTERNAL_ID"]: self.cb_sku_dest.set(c)
            if val == "TITLE": self.cb_title_dest.set(c)
        self.check_ready()

    def check_ready(self):
        if self.frame_origen.file_path and self.frame_destino.file_path:
            self.btn_run.state(["!disabled"])

    def log(self, msg):
        self.txt_log.insert(tk.END, f"> {msg}\n")
        self.txt_log.see(tk.END)

    def iniciar_flujo(self):
        hojas_seleccionadas = self.frame_origen.get_selected_sheets()
        if not hojas_seleccionadas:
            messagebox.showwarning("Atención", "Selecciona al menos una hoja de la lista de Origen.")
            return
        if not self.cb_sku_dest.get():
            messagebox.showwarning("Faltan Datos", "Configura las columnas de Destino (Mercado Libre).")
            return

        config_map = {}
        for hoja in hojas_seleccionadas:
            dialog = ConfigHojaDialog(self.root, self.frame_origen.file_path, hoja)
            self.root.wait_window(dialog)
            if dialog.result: config_map[hoja] = dialog.result
            else: self.log(f"Cancelado hoja '{hoja}'.")
        
        if not config_map: return

        save_path = filedialog.asksaveasfilename(defaultextension=".xlsx", 
                                                 filetypes=[("Excel", "*.xlsx")],
                                                 initialfile=f"ML_Actualizado.xlsx")
        
        try:
            if os.path.exists(save_path):
                f = open(save_path, "r+")
                f.close()
        except PermissionError:
            messagebox.showerror("Error", f"El archivo de destino está ABIERTO.\nCierra '{os.path.basename(save_path)}' y reintenta.")
            return
        except: pass

        if not save_path: return
        
        self.btn_run.config(text="⏳ PROCESANDO...", state="disabled")
        self.progress_var.set(0)
        self.lbl_progress.config(text="Iniciando...")
        threading.Thread(target=self.ejecutar_logica, args=(save_path, config_map)).start()

    def limpiar_sku_seguro(self, valor):
        if valor is None: return None
        s = str(valor).strip()
        if s.startswith("'"): s = s[1:]
        return s.strip().upper()

    def ejecutar_logica(self, save_path, config_map):
        try:
            self.log("--- INICIANDO ---")
            
            # 1. CONSOLIDAR
            self.log(f"Consolidando {len(config_map)} hojas...")
            mapa_precios = {}       
            mapa_nombres = {} # Para el reporte simple: SKU -> Nombre
            lista_nombres_master = {} 
            
            total_hojas = len(config_map)
            
            for i, (nombre_hoja, columnas) in enumerate(config_map.items()):
                self.lbl_progress.config(text=f"Leyendo: {nombre_hoja}")
                
                wb_temp = openpyxl.load_workbook(self.frame_origen.file_path, read_only=True, data_only=True)
                ws = wb_temp[nombre_hoja]
                
                sel_sku = columnas["sku"]
                sel_price = columnas["price"]
                sel_name = columnas["name"]
                
                idx_sku = -1
                idx_price = -1
                idx_name = -1
                header_idx = -1
                
                rows_iter = ws.iter_rows(values_only=True)
                for r_i, row in enumerate(rows_iter):
                    row_cols = []
                    for c_i, val in enumerate(row):
                        v_str = str(val).strip() if val else ""
                        if v_str == "": v_str = f"Columna {c_i+1}"
                        row_cols.append(v_str)
                    
                    if sel_sku in row_cols and sel_price in row_cols:
                        header_idx = r_i
                        idx_sku = row_cols.index(sel_sku)
                        idx_price = row_cols.index(sel_price)
                        if sel_name and sel_name in row_cols:
                            idx_name = row_cols.index(sel_name)
                        break
                
                if idx_sku == -1:
                    wb_temp.close()
                    continue

                wb_temp.close()
                wb_temp = openpyxl.load_workbook(self.frame_origen.file_path, read_only=True, data_only=True)
                ws = wb_temp[nombre_hoja]
                
                for r_i, row in enumerate(ws.iter_rows(values_only=True)):
                    if r_i <= header_idx: continue
                    try:
                        val_sku = row[idx_sku]
                        val_price = row[idx_price]
                        val_name = row[idx_name] if idx_name != -1 else ""
                        
                        sku_clean = self.limpiar_sku_seguro(val_sku)
                        if sku_clean:
                            if val_name:
                                nombre_limpio = str(val_name).strip()
                                lista_nombres_master[nombre_limpio] = sku_clean
                                mapa_nombres[sku_clean] = nombre_limpio # Guardamos nombre para reporte
                            
                            try:
                                price_clean = float(val_price)
                                # Guardamos tupla: (Precio, HojaOrigen)
                                mapa_precios[sku_clean] = (price_clean, nombre_hoja)
                            except: pass
                    except IndexError: pass
                
                wb_temp.close()
                self.progress_var.set(((i+1)/total_hojas)*30)

            self.log(f"Datos consolidados: {len(mapa_precios)} productos.")

            # 2. DESTINO (ESCRITURA QUIRÚRGICA)
            self.log("Actualizando Mercado Libre...")
            shutil.copy(self.frame_destino.file_path, save_path)
            
            wb_dest = openpyxl.load_workbook(save_path) 
            ws_dest = wb_dest[self.frame_destino.combo_sheet.get()]
            
            sel_sku_dest = self.cb_sku_dest.get()
            sel_title_dest = self.cb_title_dest.get()
            
            idx_sku_dest = -1
            idx_price_dest = -1
            idx_title_dest = -1
            start_row = -1
            
            for r in range(1, 20):
                row_cells = ws_dest[r]
                for c_i, cell in enumerate(row_cells):
                    val = str(cell.value).strip() if cell.value else ""
                    if val == "": val = f"Columna {c_i+1}"
                    if val == sel_sku_dest: idx_sku_dest = cell.column
                    if val == "PRICE": idx_price_dest = cell.column
                    if val == sel_title_dest: idx_title_dest = cell.column
                
                if idx_sku_dest != -1 and idx_price_dest != -1:
                    start_row = r + 1
                    break
            
            if idx_sku_dest == -1: raise Exception("Error columnas destino.")

            cambios = 0
            amarillo = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
            skus_ml_found = set()
            no_sku_list = []
            
            total = ws_dest.max_row
            
            for i, row in enumerate(range(start_row, total + 1)):
                if i % 50 == 0:
                    self.progress_var.set(30 + (i/(total-start_row)*40))
                    self.lbl_progress.config(text=f"Procesando {i}...")
                
                c_sku = ws_dest.cell(row=row, column=idx_sku_dest)
                val_sku = self.limpiar_sku_seguro(c_sku.value)
                
                if val_sku:
                    if val_sku in mapa_precios:
                        # Mapa precios ahora tiene (precio, hoja)
                        precio_nuevo, _ = mapa_precios[val_sku]
                        cell_price = ws_dest.cell(row=row, column=idx_price_dest)
                        cell_price.value = precio_nuevo
                        for c in ws_dest[row]: c.fill = amarillo
                        cambios += 1
                        skus_ml_found.add(val_sku)
                else:
                    if idx_title_dest != -1:
                        c_title = ws_dest.cell(row=row, column=idx_title_dest)
                        t = str(c_title.value).strip() if c_title.value else ""
                        if t: no_sku_list.append(t)

            wb_dest.save(save_path)
            self.progress_var.set(75)

            # 3. PREPARAR DATOS REPORTES
            self.log("Preparando vista previa...")
            
            # Datos Faltantes Limpios
            skus_missing = [s for s in mapa_precios if s not in skus_ml_found]
            data_faltantes = []
            for s in skus_missing:
                precio, hoja = mapa_precios[s]
                nombre = mapa_nombres.get(s, "")
                data_faltantes.append([s, nombre, precio, hoja])
                
            # Datos IA Limpios
            data_ia = []
            if no_sku_list and lista_nombres_master:
                self.log("Ejecutando IA preliminar...")
                keys = list(lista_nombres_master.keys())
                for t in no_sku_list:
                    match = difflib.get_close_matches(t, keys, n=1, cutoff=0.4)
                    if match:
                        best = match[0]
                        sku_s = lista_nombres_master[best]
                        score = int(difflib.SequenceMatcher(None, t, best).ratio() * 100)
                        
                        hoja_org = "-"
                        if sku_s in mapa_precios:
                            _, hoja_org = mapa_precios[sku_s]
                            
                        data_ia.append([t, best, sku_s, f"{score}%", hoja_org])

            # 4. MOSTRAR PREVIEW (En hilo principal)
            self.root.after(0, lambda: self.mostrar_preview_y_guardar(save_path, data_faltantes, data_ia, cambios, len(skus_ml_found), total_hojas))

        except Exception as e:
            self.log(f"❌ Error: {e}")
            messagebox.showerror("Error", str(e))
            self.root.after(0, lambda: self.btn_run.config(text="🚀  CONFIGURAR Y PROCESAR", state="!disabled"))

    def mostrar_preview_y_guardar(self, save_path, data_faltantes, data_ia, cambios, unicos, total_hojas):
        # Si hay faltantes, mostrar preview
        if data_faltantes:
            cols = ["SKU", "NOMBRE", "PRECIO", "HOJA ORIGEN"]
            dialog = ReportePreviewDialog(self.root, "Reporte de Faltantes", data_faltantes, cols)
            self.root.wait_window(dialog)
            
            if not dialog.confirmado:
                self.log("Generación de reportes cancelada por usuario.")
                messagebox.showinfo("Proceso Terminado", "Se actualizó el archivo principal, pero NO se generaron reportes extra.")
                self.finalizar_ui()
                return

        # Generar Reportes Reales
        try:
            if data_faltantes:
                ruta_rep = os.path.join(os.path.dirname(save_path), "Reporte_Faltantes.xlsx")
                wb_r = openpyxl.Workbook()
                ws_r = wb_r.active
                ws_r.title = "Faltantes"
                ws_r.append(["SKU", "NOMBRE PRODUCTO", "PRECIO LISTA", "HOJA ORIGEN"])
                
                for row in data_faltantes:
                    ws_r.append(row)
                    ws_r.cell(row=ws_r.max_row, column=3).number_format = '"$"#,##0.00'
                
                estilizar_hoja_excel(ws_r)
                wb_r.save(ruta_rep)

            if data_ia:
                ruta_ia = os.path.join(os.path.dirname(save_path), "AYUDA_VINCULACION.xlsx")
                wb_ia = openpyxl.Workbook()
                ws_ia = wb_ia.active
                ws_ia.append(["TITULO ML", "COINCIDENCIA", "SKU SUGERIDO", "SIMILITUD", "HOJA ORIGEN"])
                for row in data_ia:
                    ws_ia.append(row)
                estilizar_hoja_excel(ws_ia)
                wb_ia.save(ruta_ia)
                
            dups = cambios - unicos
            msg = f"Proceso Finalizado.\n\nCambios: {cambios}\nUnicos: {unicos}\nDuplicados: {dups}\nHojas: {total_hojas}"
            messagebox.showinfo("Éxito", msg)
            
        except Exception as e:
            messagebox.showerror("Error guardando reportes", str(e))
        finally:
            self.finalizar_ui()

    def finalizar_ui(self):
        self.btn_run.config(text="🚀  CONFIGURAR Y PROCESAR", state="!disabled")
        self.progress_var.set(0)
        self.lbl_progress.config(text="Listo.")

if __name__ == "__main__":
    root = tk.Tk()
    app = ActualizadorApp(root)
    # Estilos zebra para treeviews
    app.frame_destino.tree.tag_configure('odd', background='white')
    app.frame_destino.tree.tag_configure('even', background='#f4f6f9')
    root.mainloop()